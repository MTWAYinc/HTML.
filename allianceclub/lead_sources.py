"""Lectura y filtrado de leads desde los dos excels de salida de ScoutClub."""
from __future__ import annotations

import re
from urllib.parse import urlparse

import openpyxl

from models import AllianceLead

NUEVOS_LEADS_SHEET = "Nuevos Leads ScoutClub"
REFERENCIA_SHEET = "Referencia 522 (ya contactados)"
EXAMPLE_ROW_COMPANY_NAME = "Example Brand Co."

ESTADO_NUEVO = "Nuevo, no contactado"
CANAL_ALLIANCECLUB = "AllianceClub"
VEREDICTOS_VALIDOS = {"ICP directo", "ICP probable"}


def normalize_linkedin_url(url_or_handle: str) -> str:
    """Minusculas, sin esquema, sin www., sin query/fragment, sin barra final."""
    if not url_or_handle:
        return ""
    value = url_or_handle.strip().lower()
    if "://" not in value:
        value = "http://" + value
    parsed = urlparse(value)
    host = re.sub(r"^www\.", "", parsed.netloc)
    path = parsed.path.rstrip("/")
    if not host:
        return ""
    return f"{host}{path}"


def _row_dict(headers: list, row: tuple) -> dict:
    return dict(zip(headers, row))


def load_nuevos_leads(path: str) -> list[AllianceLead]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[NUEVOS_LEADS_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = list(rows[0])

    leads: list[AllianceLead] = []
    for row in rows[1:]:
        if row[0] is None or row[0] == EXAMPLE_ROW_COMPANY_NAME:
            continue
        d = _row_dict(headers, row)
        estado = (d.get("Estado") or "").strip()
        canal = d.get("Canal recomendado") or ""
        if estado != ESTADO_NUEVO:
            continue
        if CANAL_ALLIANCECLUB not in canal:
            continue
        if not d.get("LinkedIn Url"):
            continue

        leads.append(
            AllianceLead(
                company_name=str(d.get("Company Name") or ""),
                first_name=str(d.get("First Name") or ""),
                last_name=str(d.get("Last Name") or ""),
                title=str(d.get("Title") or ""),
                email=str(d.get("Email") or ""),
                linkedin_url=str(d.get("LinkedIn Url") or ""),
                website=str(d.get("Website") or ""),
                country=str(d.get("Country") or ""),
                categoria_mtway=str(d.get("Categoria MTWAY") or ""),
                source="nuevos_leads",
                score_icp=d.get("Score ICP (1-24)"),
                meta_ads_activo=str(d.get("Meta Ads Activo") or ""),
                aov_estimado=str(d.get("AOV Estimado") or ""),
            )
        )
    return leads


def load_referencia_leads(path: str) -> list[AllianceLead]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[REFERENCIA_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = list(rows[0])

    leads: list[AllianceLead] = []
    for row in rows[1:]:
        if row[0] is None:
            continue
        d = _row_dict(headers, row)
        veredicto = (d.get("Veredicto ICP") or "").strip()
        if veredicto not in VEREDICTOS_VALIDOS:
            continue
        linkedin_url = d.get("Person Linkedin Url")
        if not linkedin_url:
            continue

        leads.append(
            AllianceLead(
                company_name=str(d.get("Company Name") or ""),
                first_name=str(d.get("First Name") or ""),
                last_name=str(d.get("Last Name") or ""),
                title=str(d.get("Title") or ""),
                email=str(d.get("Email") or ""),
                linkedin_url=str(linkedin_url or ""),
                website=str(d.get("Website") or ""),
                country=str(d.get("Country") or ""),
                categoria_mtway=str(d.get("Categoria MTWAY") or ""),
                source="referencia_522",
                veredicto_icp=veredicto,
            )
        )
    return leads


def load_all_candidates(nuevos_path: str, referencia_path: str) -> list[AllianceLead]:
    return load_nuevos_leads(nuevos_path) + load_referencia_leads(referencia_path)
